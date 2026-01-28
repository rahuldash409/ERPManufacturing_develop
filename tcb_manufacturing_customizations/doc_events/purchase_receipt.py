import frappe
from openpyxl import load_workbook
import time
from frappe import _
from frappe.utils import flt

@frappe.whitelist()
def import_sap_to_purchase_receipt(file_url,supplier_Val, purchase_order=None):
    try:
        
        # comnmented to now use without saving the document
        # pr_doc = frappe.get_doc("Purchase Receipt", purchase_receipt_name)
        
        # if not pr_doc.has_permission("write"):
        #     frappe.throw(_("No permission to modify Purchase Receipt"))
        
        supplier = supplier_Val
        if not supplier:
            frappe.throw("Please select a Supplier first before importing sap file")
        
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = file_doc.get_full_path()
        
        # Load workbook with openpyxl
        workbook = load_workbook(filename=file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
        
        # Find table sections with "SR.NO" header
        table_sections = []
        for idx, row in enumerate(data):
            first_col = str(row[0]).strip() if row[0] is not None else ""
            
            if first_col == "SR.NO":
                # Check if row has "NO OF ROLLS"
                row_has_no_of_rolls = any(
                    (str(value).strip().upper() == "NO OF ROLLS")
                    for value in row if value is not None
                )
                
                if not row_has_no_of_rolls:
                    table_sections.append(idx)
        
        if not table_sections:
            frappe.throw("No valid SAP table headers found in the file")
        
        items_data = []
        total_items = 0
        skipped_items = []
        
        for section_idx, start_row in enumerate(table_sections):
            # Find end of table
            end_row = None
            for idx in range(start_row + 1, len(data)):
                if data[idx][0] is None or str(data[idx][0]).strip() == '' or \
                   (len(data[idx]) > 1 and str(data[idx][1]).strip().upper() == 'TOTAL'):
                    end_row = idx
                    break
            
            if end_row is None:
                end_row = len(data)
            
            # Get header row to map column names
            header_row = data[start_row]
            headers = {str(val).strip(): idx for idx, val in enumerate(header_row) if val is not None}
            
            # Process data rows
            for row_idx in range(start_row + 1, end_row):
                row = data[row_idx]
                
                # Skip invalid rows
                try:
                    sr_no = int(row[headers.get('SR.NO', 0)]) if headers.get('SR.NO') is not None else 0
                except (ValueError, TypeError, IndexError):
                    continue
                
                # Get Material column
                material_col_idx = headers.get('Material', 1)
                if material_col_idx >= len(row) or row[material_col_idx] is None or str(row[material_col_idx]).strip() == '':
                    continue
                
                supplier_part_no = str(row[material_col_idx]).strip()
                item_code = get_item_from_supplier_part_no(supplier, supplier_part_no)
                
                if not item_code:
                    desc_col_idx = headers.get('Material Description', 2)
                    description = str(row[desc_col_idx]).strip() if desc_col_idx < len(row) and row[desc_col_idx] else ''
                    skipped_items.append({
                        'supplier_part_no': supplier_part_no,
                        'description': description
                    })
                    continue
                
                item_details = frappe.db.get_value(
                    "Item",
                    item_code,
                    ["item_name", "stock_uom", "description"],
                    as_dict=True
                )
                
                # Get quantity and batch
                mtr_col_idx = headers.get('MTR', 3)
                batch_col_idx = headers.get('Batch', 4)
                desc_col_idx = headers.get('Material Description', 2)
                
                qty = float(row[mtr_col_idx]) if mtr_col_idx < len(row) and row[mtr_col_idx] else 0
                batch = str(row[batch_col_idx]).strip() if batch_col_idx < len(row) and row[batch_col_idx] else ''
                description = str(row[desc_col_idx]).strip() if desc_col_idx < len(row) and row[desc_col_idx] else ''
                item_price = frappe.db.get_value("Item Price",filters={"item_code":item_code,"buying":1},fieldname = ["price_list_rate"], order_by= "creation desc")

      
                item_data = {
                    'item_code': item_code,
                    'item_name': item_details.get('item_name') if item_details else description,
                    'supplier_part_no': supplier_part_no,
                    'qty': qty,
                    'uom': item_details.get('stock_uom') if item_details else 'Meter',
                    'description': item_details.get('description') if item_details else description,
                    'custom_sub_batch': batch,
                    'rate':item_price or 0
                }

                if purchase_order:
                    # Link to Purchase Order
                    po_doc = frappe.get_doc("Purchase Order", purchase_order)
                    poitem = next((it for it in po_doc.items if it.item_code == item_code), None)
                    if poitem:
                        item_data['purchase_order'] = purchase_order
                        item_data['purchase_order_item'] = poitem.name
                
                items_data.append(item_data)
                total_items += 1
        
        workbook.close()
        
        return {
            "status": "success",
            "items_count": total_items,
            "items_data": items_data,
            "skipped_count": len(skipped_items)
        }
        
    except Exception as e:
        frappe.throw("Error importing SAP file: {0}").format(str(e))

def get_item_from_supplier_part_no(supplier, supplier_part_no):
    try:
        item_code = frappe.db.get_value(
            "Item Supplier",
            {
                "supplier": supplier,
                "supplier_part_no": supplier_part_no
            },
            "parent"
        )
        return item_code
    except Exception as e:
        frappe.log_error(f"Error fetching item for supplier part no {supplier_part_no}: {str(e)}")
        return None

@frappe.whitelist()
def update_batches(docname, method=None):
    doc = frappe.get_doc("Purchase Receipt", docname)
    time.sleep(0.5)
    
    if doc.items:
        for item in doc.items:
            if item.custom_sub_batch:
                if item.serial_and_batch_bundle:
                    ba_bundle = frappe.get_doc("Serial and Batch Bundle", item.serial_and_batch_bundle)
                    for entry in ba_bundle.entries:
                        if entry.batch_no:
                            batch = frappe.get_doc("Batch", entry.batch_no)
                            if not batch.custom_sub_batch:
                                batch.custom_sub_batch = item.custom_sub_batch
                                batch.save()
                    frappe.db.commit()
            if not item.custom_sub_batch:
                if item.serial_and_batch_bundle:
                    ba_bundle = frappe.get_doc("Serial and Batch Bundle", item.serial_and_batch_bundle)
                    for entry in ba_bundle.entries:
                        if entry.batch_no:
                            batch = frappe.get_doc("Batch", entry.batch_no)
                            if not batch.custom_sub_batch:
                                batch.custom_sub_batch = batch.name
                                batch.save()
                    frappe.db.commit()

def validate(doc, method=None):
    if doc.items:
        for item in doc.items:
            if "fabric" in item.item_group.lower() and not item.custom_sub_batch:
                frappe.throw(f"Sub Batch required for item {item.item_code} in row {item.idx}")

def check_before_submit(doc, method=None):
    for item in doc.items:
        if item.custom_sub_batch:
            if frappe.db.exists("Batch", {"custom_sub_batch": item.custom_sub_batch, "item": item.item_code}):
                frappe.throw(msg=f"The Sub batch <b>{item.custom_sub_batch}</b> already exists for <b>{item.item_code}</b> at row {item.idx}. Try appending -1 to the sub batch number to create new batches.")

# TEMPORARILY HIDDEN - - DO NOT REMOVE
# @frappe.whitelist()
# def qc_check(docname, method=None):
#     doc = frappe.get_doc("Purchase Receipt", docname)
#     flag = False
#     res = ""
    
#     if doc.items:
#         for item in doc.items:
#             if item.quality_inspection:
#                 qc = frappe.get_doc("Quality Inspection", item.quality_inspection)
#                 if qc.status == "Rejected" and item.qty > item.rejected_qty:
#                     item.rejected_qty = item.qty
#                     item.qty = 0
#                     flag = True
    
#     if flag:
#         doc.save(ignore_permissions=True)
#         frappe.db.commit()
    
#     return res

@frappe.whitelist()
def set_container_data(docname, method=None):
    doc = frappe.get_doc("Purchase Receipt", docname)
    time.sleep(0.5)
    
    if doc.custom_container_name and doc.custom_container_number:
        for item in doc.items:
            if item.serial_and_batch_bundle:
                ba_bundle = frappe.get_doc("Serial and Batch Bundle", item.serial_and_batch_bundle)
                for entry in ba_bundle.entries:
                    if entry.batch_no:
                        batch = frappe.get_doc("Batch", entry.batch_no)
                        if not batch.custom_container_name and not batch.custom_container_number:
                            batch.custom_container_name = doc.custom_container_name
                            batch.custom_container_number = doc.custom_container_number
                            batch.save()
                        
                        if batch.custom_segregated_item_qty < 1:
                            batch.custom_segregated_item_qty = flt(item.custom_segregated_bags_qty) or 0
                            batch.save()

                frappe.db.commit()


def create_import_bales(purchase_receipt_doc, method=None):
    """
    Auto-create Import Bales when Purchase Receipt is submitted.

    NEW FLOW: Only creates bales for items with Item Group = "packaged ad*star bags"

    This replaces the old flow that used different item group check.
    """
    try:
        from tcb_manufacturing_customizations.bales_utils import create_bales_from_purchase_receipt

        created_bales = create_bales_from_purchase_receipt(purchase_receipt_doc)

        if created_bales:
            frappe.msgprint(
                _("Created {0} Import Bales: {1}").format(
                    len(created_bales), ", ".join(created_bales)
                ),
                indicator="green",
                alert=True,
            )
    except Exception as e:
        # Log error but don't block PR submission
        frappe.log_error(
            f"Error creating bales for PR {purchase_receipt_doc.name}: {str(e)}",
            "Import Bales Creation Error",
        )


def cancel_import_bales(purchase_receipt_doc, method=None):
    """
    Cancel/delete bales when Purchase Receipt is cancelled.
    Purchase Receipt is the SINGLE AUTHORITY for Import Bales lifecycle.

    Uses the new bales_utils for cancellation.
    """
    try:
        from tcb_manufacturing_customizations.bales_utils import cancel_bales_from_purchase_receipt

        cancel_bales_from_purchase_receipt(purchase_receipt_doc)
    except Exception as e:
        frappe.log_error(
            f"Error cancelling bales for PR {purchase_receipt_doc.name}: {str(e)}",
            "Import Bales Cancel Error",
        )


def testrate(item_code="LA-CR-NF-380"):
    item_price = frappe.db.get_value("Item Price",filters={"item_code":item_code},fieldname = ["price_list_rate"], order_by= "creation desc")
    print(item_price)
    
    
    
    
 
# def linkpo(doc,method=None):
    
#     itemnqty = {}
    
#     if doc.custom_purchase_order:
#         po_doc = frappe.get_doc("Purchase Order",doc.custom_purchase_order)
        
#         for item in doc.items:
#         # USE THIS TO GET EACH ROW OF QTY
#         # itemnqty.setdefault(item.item_code,[]).append(item.received_qty)
        
#         # THIS IS LITTLE COMPLEX, BUT WORKS WELL. THIS TOTALS THE ITEM QTY PER ITEM
#             itemnqty[item.item_code] = itemnqty.get(item.item_code,0) + item.received_qty
             
#     print(f"======================={itemnqty}=================================")
#     if itemnqty:
#         for item in itemnqty:
#             poitem = next((it for it in po_doc.items if it.item_code == item),None)
#             if poitem:
#                 itemname = poitem.name
#                 for row in doc.items:
#                     if row.item_code == item:
#                         row.purchase_order = doc.custom_purchase_order
#                         row.purchase_order_item = itemname
        
        
# REMOVE PO REFERENCE ON DELETION OF DOCUMENT
def removeporef(doc,method=None):
    if doc.custom_purchase_order:
        doc.custom_purchase_order = ""
        