import frappe
from frappe.utils import cint, flt, strip_html, getdate, today, add_days
import io


def get_stock_balance_from_sle(item_code, warehouse, posting_date):
    """Get stock balance from Stock Ledger Entry for a specific date"""
    posting_date = getdate(posting_date) if posting_date else getdate(today())

    result = frappe.db.sql("""
        SELECT qty_after_transaction
        FROM `tabStock Ledger Entry`
        WHERE item_code = %s
        AND warehouse = %s
        AND posting_date <= %s
        AND docstatus < 2
        AND is_cancelled = 0
        ORDER BY posting_datetime DESC, creation DESC
        LIMIT 1
    """, (item_code, warehouse, posting_date), as_dict=True)

    return flt(result[0].qty_after_transaction) if result else 0


def get_ordered_qty_by_date(item_code, warehouse, posting_date):
    """Get ordered qty from Purchase Orders created on or before the posting date"""
    posting_date = getdate(posting_date) if posting_date else getdate(today())

    result = frappe.db.sql("""
        SELECT SUM((poi.qty - poi.received_qty) * poi.conversion_factor) as ordered_qty
        FROM `tabPurchase Order Item` poi
        INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE poi.item_code = %s
        AND poi.warehouse = %s
        AND poi.qty > poi.received_qty
        AND po.docstatus = 1
        AND po.transaction_date <= %s
        AND po.status NOT IN ('Closed', 'Delivered')
        AND poi.delivered_by_supplier = 0
    """, (item_code, warehouse, posting_date), as_dict=True)

    return flt(result[0].ordered_qty) if result and result[0].ordered_qty else 0


def get_wo_remaining_qty_by_date(item_code, warehouse, posting_date):
    """Get Work Order remaining qty for orders created on or before the posting date"""
    posting_date = getdate(posting_date) if posting_date else getdate(today())

    result = frappe.db.sql("""
        SELECT SUM(woi.required_qty - woi.consumed_qty) as wo_remaining_qty
        FROM `tabWork Order Item` woi
        INNER JOIN `tabWork Order` wo ON wo.name = woi.parent
        WHERE woi.item_code = %s
        AND woi.source_warehouse = %s
        AND wo.docstatus = 1
        AND DATE(wo.creation) <= %s
        AND woi.required_qty > woi.consumed_qty
    """, (item_code, warehouse, posting_date), as_dict=True)

    return flt(result[0].wo_remaining_qty) if result and result[0].wo_remaining_qty else 0


@frappe.whitelist()
def get_dashboard_stats(warehouse=None, posting_date=None):
    """Get dashboard statistics for raw materials"""
    posting_date = getdate(posting_date) if posting_date else getdate(today())

    # Exclude Segregation Warehouse and Finished Goods warehouses
    conditions = "w.warehouse_name NOT LIKE %s AND w.warehouse_name NOT LIKE %s"
    params = ["%Segregation%", "%Finished Goods%"]

    if warehouse:
        conditions += " AND b.warehouse = %s"
        params.append(warehouse)

    # Get all item-warehouse combinations
    data = frappe.db.sql("""
        SELECT
            b.item_code,
            b.warehouse,
            i.item_name
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        INNER JOIN `tabWarehouse` w ON w.name = b.warehouse
        WHERE {conditions}
        ORDER BY b.item_code, b.warehouse
    """.format(conditions=conditions), tuple(params), as_dict=True)

    # Calculate qty values for each row
    total_actual_qty = 0
    total_po_qty = 0
    total_wo_remaining = 0
    total_available_qty = 0
    in_stock_count = 0
    out_of_stock_count = 0
    items_with_qty = []

    for row in data:
        item = row.get('item_code')
        wh = row.get('warehouse')
        actual_qty = get_stock_balance_from_sle(item, wh, posting_date)
        po_qty = get_ordered_qty_by_date(item, wh, posting_date)
        wo_remaining_qty = get_wo_remaining_qty_by_date(item, wh, posting_date)
        available_qty = actual_qty + po_qty - wo_remaining_qty

        total_actual_qty += actual_qty
        total_po_qty += po_qty
        total_wo_remaining += wo_remaining_qty
        total_available_qty += available_qty

        if available_qty > 0:
            in_stock_count += 1
        else:
            out_of_stock_count += 1

        items_with_qty.append({
            'item_code': item,
            'item_name': row.get('item_name'),
            'warehouse': wh,
            'available_qty': available_qty
        })

    # Get top 10 shortage items (lowest available qty)
    top_shortage = sorted(items_with_qty, key=lambda x: x['available_qty'])[:10]

    # Get trend data for last 7 days
    trend_data = []
    for i in range(6, -1, -1):
        date = add_days(posting_date, -i)
        day_actual = 0
        day_po = 0
        day_wo = 0

        for row in data:
            item = row.get('item_code')
            wh = row.get('warehouse')
            day_actual += get_stock_balance_from_sle(item, wh, date)
            day_po += get_ordered_qty_by_date(item, wh, date)
            day_wo += get_wo_remaining_qty_by_date(item, wh, date)

        trend_data.append({
            'date': str(date),
            'actual_qty': day_actual,
            'po_qty': day_po,
            'wo_remaining': day_wo
        })

    return {
        "summary": {
            "total_items": len(data),
            "in_stock": in_stock_count,
            "out_of_stock": out_of_stock_count,
            "total_actual_qty": total_actual_qty,
            "total_available_qty": total_available_qty,
            "total_wo_remaining": total_wo_remaining
        },
        "stock_status": [
            {"status": "In Stock", "count": in_stock_count},
            {"status": "Out of Stock", "count": out_of_stock_count}
        ],
        "top_shortage": top_shortage,
        "trend_data": trend_data
    }


@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_raw_material_items(doctype, txt, searchfield, start, page_len, filters):
    """Get items that exist in warehouse bins (excluding Segregation and Finished Goods)"""
    return frappe.db.sql("""
        SELECT DISTINCT b.item_code, i.item_name
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        INNER JOIN `tabWarehouse` w ON w.name = b.warehouse
        WHERE w.warehouse_name NOT LIKE %s
        AND w.warehouse_name NOT LIKE %s
        AND (b.item_code LIKE %s OR i.item_name LIKE %s)
        ORDER BY b.item_code
        LIMIT %s OFFSET %s
    """, ("%Segregation%", "%Finished Goods%", f"%{txt}%", f"%{txt}%", cint(page_len), cint(start)))


@frappe.whitelist()
def get_raw_materials(
    item_code=None,
    item_name=None,
    item_group=None,
    warehouse=None,
    posting_date=None,
    limit_start=0,
    limit_page_length=20,
    out_of_stock_only=0,
):
    """Get raw materials from Bin (excluding Segregation and Finished Goods warehouses)"""

    posting_date = posting_date or today()
    out_of_stock_only = cint(out_of_stock_only)

    # Exclude Segregation Warehouse and Finished Goods warehouses
    conditions = "w.warehouse_name NOT LIKE %s AND w.warehouse_name NOT LIKE %s"
    params = ["%Segregation%", "%Finished Goods%"]

    if item_code:
        conditions += " AND b.item_code = %s"
        params.append(item_code)

    if item_name:
        conditions += " AND i.item_name LIKE %s"
        params.append(f"%{item_name}%")

    if item_group:
        conditions += " AND i.item_group = %s"
        params.append(item_group)

    if warehouse:
        conditions += " AND b.warehouse = %s"
        params.append(warehouse)

    # Get all data first to calculate available_qty for filtering
    all_data = frappe.db.sql(
        """
        SELECT
            b.item_code,
            b.warehouse,
            i.item_name,
            i.item_group,
            i.description,
            i.stock_uom as uom
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        INNER JOIN `tabWarehouse` w ON w.name = b.warehouse
        WHERE {conditions}
        ORDER BY b.item_code, b.warehouse
    """.format(
            conditions=conditions
        ),
        tuple(params),
        as_dict=True,
    )

    # Calculate qty values and filter based on out_of_stock_only
    filtered_data = []
    for row in all_data:
        row["description"] = strip_html(row.get("description") or "")
        item = row.get("item_code")
        wh = row.get("warehouse")

        # Get actual qty from Stock Ledger Entry for the posting date
        row["actual_qty"] = get_stock_balance_from_sle(item, wh, posting_date)
        # Get PO qty (ordered qty) for orders created on or before posting date
        row["po_qty"] = get_ordered_qty_by_date(item, wh, posting_date)
        # Get WO remaining qty for orders created on or before posting date
        row["wo_remaining_qty"] = get_wo_remaining_qty_by_date(item, wh, posting_date)
        # Calculate available qty: actual_qty + po_qty - wo_remaining_qty
        row["available_qty"] = (
            row["actual_qty"] + row["po_qty"] - row["wo_remaining_qty"]
        )

        # Apply out of stock filter
        if out_of_stock_only:
            if row["available_qty"] <= 0:
                filtered_data.append(row)
        else:
            filtered_data.append(row)

    # Get total count after filtering
    total_count = len(filtered_data)

    # Apply pagination
    start = cint(limit_start)
    end = start + cint(limit_page_length)
    paginated_data = filtered_data[start:end]

    return {"data": paginated_data, "total_count": total_count}


@frappe.whitelist()
def export_to_excel(item_code=None, item_name=None, item_group=None, warehouse=None, posting_date=None):
    """Export raw materials data to Excel with color coding"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    posting_date = posting_date or today()

    # Exclude Segregation Warehouse and Finished Goods warehouses
    conditions = "w.warehouse_name NOT LIKE %s AND w.warehouse_name NOT LIKE %s"
    params = ["%Segregation%", "%Finished Goods%"]

    if item_code:
        conditions += " AND b.item_code = %s"
        params.append(item_code)

    if item_name:
        conditions += " AND i.item_name LIKE %s"
        params.append(f"%{item_name}%")

    if item_group:
        conditions += " AND i.item_group = %s"
        params.append(item_group)

    if warehouse:
        conditions += " AND b.warehouse = %s"
        params.append(warehouse)

    # Get all data (no pagination for export)
    data = frappe.db.sql("""
        SELECT
            b.item_code,
            b.warehouse,
            i.item_name,
            i.item_group,
            i.description,
            i.stock_uom as uom
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        INNER JOIN `tabWarehouse` w ON w.name = b.warehouse
        WHERE {conditions}
        ORDER BY b.item_code, b.warehouse
    """.format(conditions=conditions), tuple(params), as_dict=True)

    # Calculate qty values for each row based on posting date
    for row in data:
        item = row.get('item_code')
        wh = row.get('warehouse')
        row['actual_qty'] = get_stock_balance_from_sle(item, wh, posting_date)
        row['po_qty'] = get_ordered_qty_by_date(item, wh, posting_date)
        row['wo_remaining_qty'] = get_wo_remaining_qty_by_date(item, wh, posting_date)
        row['available_qty'] = row['actual_qty'] + row['po_qty'] - row['wo_remaining_qty']

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Raw Materials'

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Color fills
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Write headers
    headers = ['Item Code', 'Item Name', 'Item Group', 'Warehouse', 'Description', 'UOM', 'Actual Qty', 'PO Qty', 'WO Remaining Qty', 'Available Qty']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    # Set column widths
    ws.column_dimensions['A'].width = 20  # Item Code
    ws.column_dimensions['B'].width = 30  # Item Name
    ws.column_dimensions['C'].width = 25  # Item Group
    ws.column_dimensions['D'].width = 30  # Warehouse
    ws.column_dimensions['E'].width = 40  # Description
    ws.column_dimensions['F'].width = 10  # UOM
    ws.column_dimensions['G'].width = 15  # Actual Qty
    ws.column_dimensions['H'].width = 15  # PO Qty
    ws.column_dimensions['I'].width = 18  # WO Remaining Qty
    ws.column_dimensions['J'].width = 15  # Available Qty

    # Write data with color coding
    for row_num, row_data in enumerate(data, start=2):
        available_qty = row_data.get('available_qty') or 0

        # Select fill based on available qty
        if available_qty <= 0:
            fill = red_fill
        else:
            fill = green_fill

        # Write row data (strip HTML from description for Excel)
        row_values = [
            row_data.get('item_code') or '',
            row_data.get('item_name') or '',
            row_data.get('item_group') or '',
            row_data.get('warehouse') or '',
            strip_html(row_data.get('description') or ''),
            row_data.get('uom') or '',
            row_data.get('actual_qty') or 0,
            row_data.get('po_qty') or 0,
            row_data.get('wo_remaining_qty') or 0,
            row_data.get('available_qty') or 0
        ]

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.border = thin_border
            if col >= 7:  # Numeric columns (now starts from column 7)
                cell.alignment = right_align
                cell.number_format = '#,##0.00'

    # Add legend at the bottom
    legend_row = len(data) + 4
    ws.cell(row=legend_row, column=1, value='Legend:').font = Font(bold=True)

    legend_items = [
        (legend_row + 1, 'In Stock (> 0)', green_fill),
        (legend_row + 2, 'Out of Stock (≤ 0)', red_fill)
    ]

    for row, text, fill in legend_items:
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = fill
        cell.border = thin_border

    # Save file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"Raw_Material_Dashboard_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Save to File doctype for Frappe Cloud compatibility
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "content": output.getvalue(),
        "is_private": 1
    })
    file_doc.save(ignore_permissions=True)

    return {"file_url": file_doc.file_url}
